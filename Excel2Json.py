# -*- coding: utf-8 -*-

import json

from enum import Enum
import pathlib

import openpyxl

TO_MM = 20 / 0.7875

def _setPrintArea(ws, wsJson):
    assert ws.print_area, "印刷範囲を設定して下さい. (sheet : %s)" % ws.title
    wsJson['print_area'] = ws.print_area[0]

_PAGE_SETUP_LIST = [
    'orientation', 'paperSize', 
    # 以下は、LibreOfficeで作成したシートを読込むときに下記エラーが発生する.
    # [error]AttributeError: 'NoneType' object has no attribute 'sheet_properties'
    # 注意：以下の全てのプロパティを確認してはいない。
    # ---------------------
    #'autoPageBreaks', 'blackAndWhite', 'cellComments', 'copies', 'draft',
    #'firstPageNumber', 'fitToHeight', 'fitToPage', 'fitToWidth', 'horizontalDpi',
    #'orientation', 'pageOrder', 'paperHeight', 'paperSize', 'paperWidth',
    #'scale', 'useFirstPageNumber', 'usePrinterDefaults', 'verticalDpi',
    #'sheet_properties',
]

_PAGE_SETUP_PROPERTIES_LIST = [
    'autoPageBreaks', 'fitToPage', 
]

_PAGE_MARGIN_LIST = ['top', 'bottom', 'left', 'right']

_CELL_ALIGNMENT_LIST = [
    'horizontal', 'indent', 'justifyLastLine', 'readingOrder', 'relativeIndent',
    'shrinkToFit', 'shrink_to_fit', 'textRotation', 'text_rotation', 'vertical',
    'wrapText', 'wrap_text'
]

_CELL_FONT_ATTRIBUTE_LIST = [
    'UNDERLINE_DOUBLE', 'UNDERLINE_DOUBLE_ACCOUNTING', 'UNDERLINE_SINGLE',
    'UNDERLINE_SINGLE_ACCOUNTING', 'bold', 'charset', 'color', 'condense',
    'extend', 'family', 'italic', 'name', 'outline', 'scheme', 'shadow',
    'size', 'strike', 'strikethrough', 'underline', 'vertAlign'
]

def _editColorAttribute(c):
    '''
    色属性の編集を行う
    openpyxl.styles.colors.Color クラスのコンストラクタから引用して作成
    '''
    a = {}
    if c.index is not None:
        a['indexed'] = c.index
    if c.indexed is not None:
        a['type'] = 'indexed'
        a['indexed'] = c.indexed.name
    elif c.theme is not None:
        a['type'] = 'theme'
        a['theme'] = c.theme
    elif c.auto is not None:
        a['type'] = 'auto'
        a['auto'] = c.auto
    else:
        a['type'] = 'rbg'
        a['rbg'] = c.rbg
    a['tint'] = c.tint
    return a

def _getProperties(o, l):
    p = {}
    for item in l:
        v = getattr(o, item)
        if isinstance(v, openpyxl.styles.colors.Color):
            v = _editColorAttribute(v)
        elif isinstance(v, openpyxl.worksheet.properties.PageSetupProperties):
            v = _getProperties(v, _PAGE_SETUP_PROPERTIES_LIST)
        p[item] = v
    return p

def _mergedCell(ws, cell):
    '''
    指定のセルが結合セルのどの位置に該当するかを調べる
    戻り値
     (1) 左上のセルの場合、結合セルの範囲を返す（例：B5:E6）
     (2) 接合セル内に存在するが、左上でない場合、True
     (3) 指定のセルが結合セルでない場合、None
    '''
    row = cell.row
    col = cell.column
    for r in ws.merged_cell_ranges:
        if r.bounds[0] == col and r.bounds[1] == row:
            # 結合セルの左上と同じ位置
            return r.coord
        if r.bounds[0] > col or r.bounds[1] > row:
            # 結合セルの左または上に位置する
            continue
        if r.bounds[2] < col or r.bounds[3] < row:
            # 結合セルの右または下に位置する
            continue
        # 結合セルのいづれかに位置する
        return True
    return None

def _mergedCellPosition(ws, cell):
    '''
    指定のセルが結合セルのどの位置に接しているかを調べる
    戻り値
        r['left']   : 左辺が結合セルに接している場合、True
        r['top']    : 上部が結合セルに接している場合、True
        r['right']  : 右辺が結合セルに接している場合、True
        r['bottom'] : 底辺が結合セルに接している場合、True
    '''
    row = cell.row
    col = cell.column
    for r in ws.merged_cell_ranges:
        if r.bounds[0] == col and r.bounds[1] == row:
            # 結合セルの左上と同じ位置
            return {'left': True, 'top': True, 'right': col == r.bounds[2], 'bottom': row == r.bounds[3]}
        if r.bounds[0] > col or r.bounds[1] > row:
            # 結合セルの左または上に位置する
            continue
        if r.bounds[2] < col or r.bounds[3] < row:
            # 結合セルの右または下に位置する
            continue
        # 結合セルのいづれかに位置する
        return {'left'  : col == r.bounds[0]
              , 'top'   : row == r.bounds[1]
              , 'right' : col == r.bounds[2]
              , 'bottom': row == r.bounds[3]}
    # どの結合セルに接していない
    return {'left': True, 'top': True, 'right': True, 'bottom': True}

def _editCellJson(cell, a1):
    c = {}
    c['A1'] = a1
    c['value'] = cell.value
    c['alignment'] = _getProperties(cell.alignment, _CELL_ALIGNMENT_LIST)
    c['font'] = _getProperties(cell.font, _CELL_FONT_ATTRIBUTE_LIST)
    return c

class _BOARDER_TYPE(Enum):
                      # LTRB(Left, Top, Right, Bottom)
    CELL_BOX         = 'TTTT'   # 1セルで完結
    ROW_BOX          = 'TT_T'   # 1行で矩形を描く（開始）、終了は「BOX_RIGHT_BOTTOM」
    BOTTOM_ONLY      = '___T'   # 下線のみ
    TOP_ONLY         = '_T__'   # 上線のみ
    TOP_BOTTOM       = '_T_T'   # 上下線のみ
    BOX_LEFT_TOP     = 'TT'     # 複数の行列で矩形を描く時の左上
    BOX_RIGHT_BOTTOM = '__TT'   # 複数の行列で矩形を描く時の右下

def _borderKind(b, r):
    ltrb  = 'T' if r['left'] and b.left.style else '_'
    ltrb += 'T' if r['top']  and b.top.style else '_'
    if ltrb == _BOARDER_TYPE.BOX_LEFT_TOP.value:
        return _BOARDER_TYPE.BOX_LEFT_TOP
    ltrb += 'T' if r['right']  and b.right.style else '_'
    ltrb += 'T' if r['bottom'] and b.bottom.style else '_'
    if ltrb == _BOARDER_TYPE.CELL_BOX.value:
        return _BOARDER_TYPE.CELL_BOX
    if ltrb == _BOARDER_TYPE.ROW_BOX.value:
        return _BOARDER_TYPE.ROW_BOX
    if ltrb == _BOARDER_TYPE.BOTTOM_ONLY.value:
        return _BOARDER_TYPE.BOTTOM_ONLY
    if ltrb == _BOARDER_TYPE.TOP_ONLY.value:
        return _BOARDER_TYPE.TOP_ONLY
    if ltrb == _BOARDER_TYPE.TOP_BOTTOM.value:
        return _BOARDER_TYPE.TOP_BOTTOM
    if ltrb == _BOARDER_TYPE.BOX_RIGHT_BOTTOM.value:
        return _BOARDER_TYPE.BOX_RIGHT_BOTTOM
    return None

def _getCells(ws, r):
    l = []
    for row in ws[r]:
        for cell in row:
            if not cell.value:
                continue
            rng = _mergedCell(ws, cell)
            if isinstance(rng, str):
                # 結合セルの右上
                l.append(_editCellJson(cell, rng))
            elif isinstance(rng, bool):
                # 結合セルの右上以外
                continue
            else:
                l.append(_editCellJson(cell, cell.coordinate))
    return l

def _findBoarder(ws, r):

    def __endBorder(boarder):
        if prevA1 != boarder['A1']:
            boarder['A1'] = boarder['A1'] + ':' + prevA1
        boarders.append(boarder)

    def __rectBorder(boarder):
        # 左上を検出した時に呼び出され右下のセルを検出する
        # 右に進む
        _c = None
        for _cn in range(cell.column, row[-1].column + 1):
            _a1 = '%s%d' % (openpyxl.utils.get_column_letter(_cn), row[-1].row)
            mr = _mergedCellPosition(ws, ws[_a1])
            if mr['right'] and ws[_a1].border.right.style:
                _c = openpyxl.utils.get_column_letter(_cn)
                break
        assert _c, '矩形の右端が見つからない.'
        # 下方向に進む
        _pa = openpyxl.utils.range_boundaries(r)
        _r = None
        for _rn in range(cell.row, _pa[3]):
            _a1 = '%s%d' % (_c, _rn)
            mr = _mergedCellPosition(ws, ws[_a1])
            if mr['bottom'] and ws[_a1].border.bottom.style:
                _r = _rn
                break
        assert _r, '矩形の底辺が見つからない.'
        _a1 = '%s%d' % (_c, _r)
        if _a1 != boarder['A1']:
            boarder['A1'] = boarder['A1'] + ':' + _a1
        return boarder
    def __otherRect():
        for _b in boarders:
            _pa = openpyxl.utils.range_boundaries(_b['A1'])
            if _pa[0] <= cell.column and cell.column <= _pa[2] and \
               _pa[1] <= cell.row    and cell.row    <= _pa[3]:
               return True
        return False

    boarders = []
    boarder = None
    prevA1  = None
    for row in ws[r]:
        for cell in row:
            if __otherRect():
                prevA1 = cell.coordinate
                continue
            b = ws[cell.coordinate].border
            k = _borderKind(b, _mergedCellPosition(ws, cell))
            if boarder:
                if not k:
                    # 矩形または直線の終了
                    __endBorder(boarder)
                    boarder = None
            elif k == _BOARDER_TYPE.CELL_BOX:
                boarders.append({'kind': k, 'A1': cell.coordinate})
            elif k == _BOARDER_TYPE.BOX_LEFT_TOP:
                # 矩形の左上
                boarders.append(__rectBorder({'kind': k, 'A1': cell.coordinate}))
            elif k:
                # 矩形または直線の開始
                boarder = {'kind': k, 'A1': cell.coordinate}
            elif boarder:
                # 矩形または直線の終了
                __endBorder(boarder)
                boarder = None
            prevA1 = cell.coordinate
        # 1行終了
        if boarder:
            if boarder['kind'] != _BOARDER_TYPE.BOX_LEFT_TOP:
                __endBorder(boarder)
                boarder = None

    if boarders:
        return boarders
    return None

def _getCellHeightWidth(ws, r):
    width_map = {}
    for cell in ws[r][0]:
        if hasattr(cell, 'column_letter'):
            cl = cell.column_letter
        else:
            cl = openpyxl.utils.get_column_letter(cell.column)
        width_map[cl] = ws.column_dimensions[cl].width

    height_map = {}
    for row in ws[r]:
        r = row[0].row
        wk = ws.row_dimensions[r]
        #height_map[r] = wk.height if wk.height else openpyxl.utils.units.DEFAULT_ROW_HEIGHT
        height_map[r] = wk.height

    return {'width': width_map, 'height': height_map}

#        if cellWidth == 0:
#            cellWidth += 
#            width_list.append[cellWidth]
#        print(row)

def readExcel(excelPath):
    wbJson = {}
    for ws in openpyxl.load_workbook(excelPath):
        wsJson = {}
        _setPrintArea(ws, wsJson)
        wsJson['PrintPageSetup'] = _getProperties(ws.page_setup, _PAGE_SETUP_LIST)
        wsJson['page_margins'] = _getProperties(ws.page_margins, _PAGE_MARGIN_LIST)
        printArea = wsJson['print_area']
        wsJson['cells'] = _getCells(ws, printArea)
        wsJson['row_height'] = _getCellHeightWidth(ws, printArea)
        wsJson['boarders'] = _findBoarder(ws, printArea)
        wbJson[ws.title] = wsJson
    return wbJson

def jsonOut(dictVals, jsonPath):
    #print(dictVals)
    #j = json.dumps(dictVals, default=str, indent=2)
    #print(j)
    with open(jsonPath, mode='wt', encoding='utf-8') as f:
        json.dump(dictVals, f, ensure_ascii=False, indent=2, default=str)

def _on_pairs(itr):
    #print(itr)
    d = {}
    for k, v in itr:
        if isinstance(v, str):
            if v.startswith('_BOARDER_TYPE.'):
                x = v.split('.', 2)
                v = getattr(_BOARDER_TYPE, x[1])
        d[k] = v
    return d

def jsonRead(jsonPath):
    with open(jsonPath, mode='r', encoding='utf-8') as f:
        return json.load(f, object_pairs_hook=_on_pairs)

if __name__ == '__main__':
    excelFile = 'V01-frame_001_LibreOffice.xlsx'
    excelPath = pathlib.Path(__file__).parent / 'FrameExcel' / excelFile
    print(excelPath)
    wbJson = readExcel(excelPath)
    jsonPath = '.\\mydata.json'
    jsonOut(wbJson, jsonPath)
    r = jsonRead(jsonPath)
    print(r)

#[EOF]